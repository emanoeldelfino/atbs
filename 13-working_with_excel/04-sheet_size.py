import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']

num_rows = sheet.max_row
num_cols = sheet.max_column

print(f'Rows: {num_rows}')
print(f'Colums: {num_cols}')

print(f'\n{sheet.title} dimension {num_rows}x{num_cols}', end='\n\n')

