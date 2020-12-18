import openpyxl
from openpyxl.utils import get_column_letter
import sys

wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.active

num_cols = sheet.max_column
num_rows = sheet.max_row

print(end='\n  ')
for num_col in range(1, num_cols + 1):
        print(f'| {get_column_letter(num_col):^19} ', end='')
print(' |')

for num_row in range(1, num_rows):
	print(f'{num_row} ', end='')
	for cell in sheet[num_row]:
		if '.' not in str(cell.value):
			print(f'| {str(cell.value):^19} ', end='')
		else:
			print(f'| {str(cell.value).split(".")[0]} ', end='')
	print(' |')
print()
