import openpyxl
import pprint
import string
import sys

args = sys.argv[1:]

alphabet = string.ascii_uppercase

wb = openpyxl.load_workbook(args[0])
sheet = wb.active # wb['Sheet1']

table = [[] for _ in range(sheet.max_row)]

for row in range(1, sheet.max_row + 1):
    for column in range(sheet.max_column):
        table[row - 1].append(sheet[row][column].value)

for column_letter in alphabet[:len(table[0])]:
    print(f'{column_letter:^11}', end='')
print()

for row in range(sheet.max_row):
    print(row, end=' ')
    for column in range(len(table[row])):
        print(f'{table[row][column]:<11}', end='')
    print()
