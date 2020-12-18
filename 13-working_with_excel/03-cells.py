import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb['Sheet1']

print(sheet['A1'])
print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print('Row %s, Column %s is %s' % (c.row, c.column, c.value))

print('Cell %s is %s' % (c.coordinate, c.value))

print(sheet['C1'].value)

print(sheet.cell(row=1, column=2))

print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
	print(i, sheet.cell(row=i, column=2).value)

